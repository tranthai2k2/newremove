"""
CivitAI LoRA Scraper - GUI
pip install requests openpyxl
python civitai_scraper.py
"""
import re, time, threading, tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

USERNAME = "jose9967729568"
API_URL  = "https://civitai.com/api/v1/models"
META     = {'Manhua','Manhwa','Webtoon','Illustrious','Pony','SDXL','Flux','IL','PN','New'}

def is_native(s):
    return bool(re.search(r'[\u1100-\u11FF\u3130-\u318F\uAC00-\uD7AF\u4E00-\u9FFF\u3400-\u4DBF]', s))

def parse_title(name):
    parts = [p.strip() for p in name.split('|') if p.strip()]
    m = re.match(r'^(.+?)\s*[\(\（]\s*(.+?)\s*[\)\）]\s*$', parts[0])
    char_name = m.group(1).strip() if m else parts[0]
    char_alt  = m.group(2).strip() if m else ''
    content = [p for p in parts[1:] if p not in META]
    series, alts = '', []
    for p in content:
        if not is_native(p) and not series: series = p
        else: alts.append(p)
    genre = next((p for p in parts if p in ('Manhwa','Manhua')), '')
    return [char_name, char_alt, series, ' | '.join(alts), genre]

def fetch_and_save(api_key, out_path, log, prog, btn_start, total_var, parsed_var):
    headers = {"User-Agent": "Mozilla/5.0"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    models, cursor, page = [], None, 0

    while True:
        params = {'username': USERNAME, 'types': 'LORA',
                  'limit': 100, 'sort': 'Newest', 'nsfw': 'true'}
        if cursor:
            params['cursor'] = cursor
        try:
            r = requests.get(API_URL, params=params, headers=headers, timeout=30)
        except Exception as e:
            log(f"Lỗi: {e}"); break

        if r.status_code == 401:
            log("401 — API key sai hoặc thiếu quyền NSFW."); break
        if r.status_code != 200:
            log(f"HTTP {r.status_code}"); break

        data  = r.json()
        items = data.get('items', [])
        meta  = data.get('metadata', {})
        models.extend(items)
        page += 1
        total = meta.get('totalItems', 0) or 1
        pct   = len(models) / total * 100

        prog['value'] = pct
        total_var.set(len(models))
        log(f"Page {page:>3}: +{len(items):>3}  →  {len(models)}/{total}")

        cursor = meta.get('nextCursor')
        if not cursor or not items:
            break
        time.sleep(0.5)

    titles = []
    for m in models:
        name = (m.get('name') or '').strip()
        if name:
            titles.append(name)

    titles.sort(key=str.lower)
    parsed_var.set(len(titles))
    log(f"\nTitles: {len(titles)}")

    # Build xlsx
    HDR_BG='FF2D5986'; HDR_FG='FFFFFFFF'; BC='FFB8CCE4'
    thin=Side(style='thin',color=BC); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

    wb=Workbook(); ws=wb.active; ws.title='LoRA Titles'
    cell=ws.cell(row=1,column=1,value='Title')
    cell.font=Font(name='Arial',bold=True,color=HDR_FG,size=11)
    cell.fill=PatternFill('solid',start_color=HDR_BG)
    cell.alignment=Alignment(horizontal='center',vertical='center')
    cell.border=bdr
    ws.column_dimensions['A'].width=80
    ws.row_dimensions[1].height=22

    for i,title in enumerate(titles,2):
        cell=ws.cell(row=i,column=1,value=title)
        cell.font=Font(name='Arial',size=10)
        cell.alignment=Alignment(vertical='center',wrap_text=True)
        cell.border=bdr
        ws.row_dimensions[i].height=16
    ws.freeze_panes='A2'

    wb.save(out_path)
    prog['value']=100
    log(f"Saved → {out_path}")
    messagebox.showinfo("Xong!", f"Đã lưu {len(titles)} titles\n{out_path}")
    btn_start.config(state='normal')

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CivitAI LoRA Scraper")
        self.geometry("620x520")
        self.resizable(False,False)

        pad = {'padx':16,'pady':6}

        # API Key
        tk.Label(self, text="API Key (civitai.com/user/account → API Keys):").pack(anchor='w', **pad)
        self.key_var = tk.StringVar()
        e = tk.Entry(self, textvariable=self.key_var, show='*', width=72)
        e.pack(fill='x', padx=16)

        # Output path
        tk.Label(self, text="Lưu file:").pack(anchor='w', **pad)
        frm = tk.Frame(self); frm.pack(fill='x', padx=16)
        self.path_var = tk.StringVar(value="lora_characters.xlsx")
        tk.Entry(frm, textvariable=self.path_var, width=58).pack(side='left')
        tk.Button(frm, text="...", command=self.choose_path).pack(side='left', padx=4)

        # Stats
        sf = tk.Frame(self); sf.pack(fill='x', padx=16, pady=8)
        self.total_var  = tk.IntVar(value=0)
        self.parsed_var = tk.IntVar(value=0)
        for label, var in [("Đã fetch:", self.total_var), ("Đã parse:", self.parsed_var)]:
            tk.Label(sf, text=label).pack(side='left')
            tk.Label(sf, textvariable=var, font=('Arial',11,'bold'), width=6).pack(side='left', padx=(0,16))

        # Progress
        self.prog = ttk.Progressbar(self, length=588, mode='determinate')
        self.prog.pack(padx=16, pady=4)

        # Log
        tk.Label(self, text="Log:").pack(anchor='w', padx=16)
        self.log_box = tk.Text(self, height=12, font=('Courier',9), state='disabled', bg='#f8f8f8')
        self.log_box.pack(fill='both', padx=16, pady=4)

        # Button
        self.btn = tk.Button(self, text="▶  Fetch tất cả", font=('Arial',11,'bold'),
                             bg='#2D5986', fg='white', command=self.start, pady=6)
        self.btn.pack(fill='x', padx=16, pady=8)

    def choose_path(self):
        p = filedialog.asksaveasfilename(defaultextension='.xlsx',
            filetypes=[('Excel','*.xlsx')], initialfile='lora_characters.xlsx')
        if p: self.path_var.set(p)

    def log(self, msg):
        self.log_box.config(state='normal')
        self.log_box.insert('end', msg+'\n')
        self.log_box.see('end')
        self.log_box.config(state='disabled')

    def start(self):
        self.btn.config(state='disabled')
        self.log_box.config(state='normal'); self.log_box.delete('1.0','end'); self.log_box.config(state='disabled')
        self.prog['value']=0
        t = threading.Thread(target=fetch_and_save, args=(
            self.key_var.get().strip(),
            self.path_var.get(),
            self.log, self.prog, self.btn,
            self.total_var, self.parsed_var
        ), daemon=True)
        t.start()

if __name__ == '__main__':
    App().mainloop()