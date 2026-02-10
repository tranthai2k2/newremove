import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import os
import openpyxl
import json
from pathlib import Path
from datetime import datetime


class ImageCompareApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Image Compare & Note Tool - Dual Mode")
        self.root.geometry("1600x900")

        self.config_file = "image_compare_config.json"

        self.folder1_path = tk.StringVar()
        self.folder2_path = tk.StringVar()
        self.current_index = 0
        self.data_list = []
        self.txt_mode = tk.StringVar(value='all_tags')
        self.txt_file_choice = tk.StringVar(value='all_tags')
        self.images_per_row = tk.IntVar(value=1)
        self.all_tags_content = []
        self.canvas2_list = []
        self.view_mode = tk.StringVar(value='single')
        self.gallery_thumbnails = []

        self.load_config()
        self.setup_ui()
        self.bind_shortcuts()

    # ===================== HOTKEYS 1–7 + DELETE =====================
    def bind_shortcuts(self):
        # Điều hướng
        self.root.bind_all('<Left>', lambda e: (self.previous_image(), 'break')[1])
        self.root.bind_all('<Right>', lambda e: (self.next_image(), 'break')[1])
        self.root.bind_all('<Delete>', lambda e: (self.clear_note(), 'break')[1])

        # Helper để tránh lặp code, và tránh gõ số vào Text
        def note_hotkey(text):
            self.add_quick_note(text)
            return 'break'

        # Quick notes 1-7
        self.root.bind_all('1', lambda e: note_hotkey("thieu faceless male"))
        self.root.bind_all('2', lambda e: note_hotkey("thieu bald"))
        self.root.bind_all('3', lambda e: note_hotkey("khong can them fat"))
        self.root.bind_all('4', lambda e: note_hotkey("them fat man"))
        self.root.bind_all('5', lambda e: note_hotkey("lỗi có đầu cụt"))
        self.root.bind_all('6', lambda e: note_hotkey("clothed female nude male"))
        self.root.bind_all('7', lambda e: note_hotkey("khong can them faceless male + bald"))

    # ===================== CONFIG LOAD/SAVE =====================
    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.folder1_path.set(config.get('folder1', ''))
                    self.folder2_path.set(config.get('folder2', ''))
                    self.txt_mode.set(config.get('txt_mode', 'all_tags'))
                    self.txt_file_choice.set(config.get('txt_file_choice', 'all_tags'))
            except:
                pass

    def save_config(self):
        config = {
            'folder1': self.folder1_path.get(),
            'folder2': self.folder2_path.get(),
            'txt_mode': self.txt_mode.get(),
            'txt_file_choice': self.txt_file_choice.get()
        }
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

    # ===================== UI SETUP =====================
    def setup_ui(self):
        folder_frame = ttk.Frame(self.root, padding="10")
        folder_frame.pack(fill=tk.X)

        # Row 0: Folder 1 + TXT Mode Selection
        ttk.Label(folder_frame, text="Folder 1 (Anh):").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(folder_frame, textvariable=self.folder1_path, width=40).grid(row=0, column=1, padx=5)
        ttk.Button(folder_frame, text="Browse", command=lambda: self.browse_folder(1)).grid(row=0, column=2)

        # TXT Mode Selection
        mode_frame = ttk.LabelFrame(folder_frame, text="Chế độ TXT", padding="5")
        mode_frame.grid(row=0, column=3, padx=10)
        ttk.Radiobutton(
            mode_frame, text="Anh-TXT",
            variable=self.txt_mode,
            value='individual',
            command=self.on_txt_mode_change
        ).pack(anchor=tk.W)
        ttk.Radiobutton(
            mode_frame, text="out_tags folder",
            variable=self.txt_mode,
            value='all_tags',
            command=self.on_txt_mode_change
        ).pack(anchor=tk.W)

        # Row 1: Folder 2 + Images per row slider
        ttk.Label(folder_frame, text="Folder 2 (Anh):").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(folder_frame, textvariable=self.folder2_path, width=40).grid(row=1, column=1, padx=5)
        ttk.Button(folder_frame, text="Browse", command=lambda: self.browse_folder(2)).grid(row=1, column=2)

        slider_frame = ttk.Frame(folder_frame)
        slider_frame.grid(row=1, column=3, padx=10)
        ttk.Label(slider_frame, text="So anh F2:").pack(side=tk.LEFT)
        ttk.Scale(
            slider_frame, from_=1, to=4,
            variable=self.images_per_row,
            orient=tk.HORIZONTAL,
            length=100,
            command=self.on_slider_change
        ).pack(side=tk.LEFT, padx=5)
        self.slider_label = ttk.Label(slider_frame, text="1", font=('Arial', 10, 'bold'))
        self.slider_label.pack(side=tk.LEFT)

        # Row 2: Control Buttons
        control_frame = ttk.Frame(folder_frame)
        control_frame.grid(row=2, column=1, pady=10)
        ttk.Button(control_frame, text="Load Data", command=self.load_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Single View", command=lambda: self.switch_view('single')).pack(side=tk.LEFT, padx=2)
        ttk.Button(control_frame, text="Gallery View", command=lambda: self.switch_view('gallery')).pack(side=tk.LEFT, padx=2)

        # TXT File Selection
        self.txt_select_frame = ttk.LabelFrame(
            folder_frame, text="Chọn File TXT từ out_tags folder", padding="10"
        )
        self.txt_select_frame.grid(row=3, column=0, columnspan=4, pady=10, sticky="ew")

        # Quick buttons for out_tags files
        quick_buttons = [
            ('all_tags', 'all_tags'),
            ('all_tag', 'all_tag'),
            ('out_tags', 'out_tags'),
            ('addfaceless', 'addfaceless'),
            ('hop_txt', 'hop_txt')
        ]

        self.txt_buttons = []
        for display_name, value in quick_buttons:
            btn = ttk.Button(
                self.txt_select_frame,
                text=display_name,
                command=lambda v=value: self.set_txt_file_choice(v),
                width=15
            )
            btn.pack(side=tk.LEFT, padx=5, pady=5)
            self.txt_buttons.append(btn)

        # Show current selection
        self.txt_choice_label = ttk.Label(
            self.txt_select_frame,
            text="Hiện tại: all_tags.txt",
            font=('Arial', 9, 'bold'),
            foreground='darkgreen'
        )
        self.txt_choice_label.pack(side=tk.LEFT, padx=10)

        self.status_label = ttk.Label(folder_frame, text="", foreground="blue")
        self.status_label.grid(row=4, column=1, columnspan=2)

        # Show/hide txt select frame based on mode
        self.update_txt_selection_visibility()

        self.main_container = ttk.Frame(self.root, padding="10")
        self.main_container.pack(fill=tk.BOTH, expand=True)

        self.nav_frame = ttk.Frame(self.root, padding="10")
        self.nav_frame.pack(fill=tk.X)

        ttk.Button(
            self.nav_frame,
            text="<< Previous (Left)",
            command=self.previous_image,
            width=20
        ).pack(side=tk.LEFT, padx=5)

        self.counter_label = ttk.Label(self.nav_frame, text="0 / 0", font=('Arial', 11, 'bold'))
        self.counter_label.pack(side=tk.LEFT, padx=20)

        ttk.Button(
            self.nav_frame,
            text="Next >> (Right)",
            command=self.next_image,
            width=20
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            self.nav_frame,
            text="Export Excel",
            command=self.export_excel,
            width=20
        ).pack(side=tk.RIGHT, padx=5)

        self.create_single_view()

    # ===================== TXT MODE HANDLING =====================
    def on_txt_mode_change(self):
        self.update_txt_selection_visibility()
        self.save_config()

    def update_txt_selection_visibility(self):
        if self.txt_mode.get() == 'all_tags':
            self.txt_select_frame.grid()
            self.status_label.config(
                text="Chế độ: out_tags folder - Chọn file TXT bên dưới",
                foreground='blue'
            )
        else:
            self.txt_select_frame.grid_remove()
            self.status_label.config(
                text="Chế độ: Anh-TXT (mỗi ảnh có 1 file .txt cùng tên)",
                foreground='blue'
            )

    def set_txt_file_choice(self, value):
        self.txt_file_choice.set(value)
        self.save_config()

        display_text = f"Hiện tại: {value}.txt"
        self.txt_choice_label.config(text=display_text)

        self.status_label.config(text=f"✓ Đã chọn: {display_text}", foreground='green')

    # ===================== VIEW SWITCHING =====================
    def switch_view(self, mode):
        self.view_mode.set(mode)
        if mode == 'single':
            self.create_single_view()
            if self.data_list:
                self.display_current()
            self.nav_frame.pack(fill=tk.X)
        else:
            self.create_gallery_view()
            self.nav_frame.pack_forget()

    def on_slider_change(self, value):
        num = int(float(value))
        self.slider_label.config(text=str(num))
        self.images_per_row.set(num)
        if self.data_list and self.view_mode.get() == 'single':
            self.create_single_view()
            self.display_current()

    # ===================== SINGLE VIEW UI =====================
    def create_single_view(self):
        for widget in self.main_container.winfo_children():
            widget.destroy()

        self.canvas2_list = []
        num_images = self.images_per_row.get()

        if num_images == 1:
            canvas_size = 400
        elif num_images == 2:
            canvas_size = 350
        else:
            canvas_size = 280

        # Col 1
        col1_frame = ttk.LabelFrame(self.main_container, text="Folder 1 - Original", padding="10")
        col1_frame.grid(row=0, column=0, padx=5, sticky="nsew", rowspan=2)

        self.canvas1 = tk.Canvas(
            col1_frame,
            width=canvas_size,
            height=canvas_size,
            bg='#f0f0f0',
            relief=tk.SOLID,
            borderwidth=2
        )
        self.canvas1.pack()

        self.filename1_label = ttk.Label(
            col1_frame,
            text="File: -",
            foreground="navy",
            wraplength=canvas_size
        )
        self.filename1_label.pack(pady=5)

        # Col 2
        col2_container = ttk.LabelFrame(self.main_container, text="Folder 2 - Generated Images", padding="10")
        col2_container.grid(row=0, column=1, padx=5, sticky="nsew")

        if num_images <= 2:
            for i in range(num_images):
                frame = ttk.Frame(col2_container, relief=tk.SOLID, borderwidth=1)
                frame.grid(row=0, column=i, padx=5, pady=5)

                canvas = tk.Canvas(
                    frame,
                    width=canvas_size,
                    height=canvas_size,
                    bg='#f0f0f0',
                    relief=tk.SOLID,
                    borderwidth=2
                )
                canvas.pack()

                label = ttk.Label(frame, text=f"Image {i+1}", foreground="darkgreen")
                label.pack(pady=2)

                self.canvas2_list.append({'canvas': canvas, 'label': label})
        else:
            for i in range(num_images):
                row_pos = i // 2
                col_pos = i % 2

                frame = ttk.Frame(col2_container, relief=tk.SOLID, borderwidth=1)
                frame.grid(row=row_pos, column=col_pos, padx=5, pady=5)

                canvas = tk.Canvas(
                    frame,
                    width=canvas_size,
                    height=canvas_size,
                    bg='#f0f0f0',
                    relief=tk.SOLID,
                    borderwidth=2
                )
                canvas.pack()

                label = ttk.Label(frame, text=f"Image {i+1}", foreground="darkgreen")
                label.pack(pady=2)

                self.canvas2_list.append({'canvas': canvas, 'label': label})

        # Col 3
        col3_frame = ttk.LabelFrame(self.main_container, text="Information & Notes", padding="10")
        col3_frame.grid(row=0, column=2, padx=5, sticky="nsew", rowspan=2)

        ttk.Label(col3_frame, text="Prompt tu TXT:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        self.prompt_text = tk.Text(
            col3_frame,
            height=8,
            width=40,
            wrap=tk.WORD,
            relief=tk.SOLID,
            borderwidth=1
        )
        self.prompt_text.pack(fill=tk.BOTH, pady=5)

        # Quick Notes (1-7)
        quick_frame = ttk.LabelFrame(col3_frame, text="Quick Notes (Phim tat: 1-7)", padding="5")
        quick_frame.pack(fill=tk.X, pady=5)

        quick_notes = [
            ("thieu faceless male", "1"),
            ("thieu bald", "2"),
            ("khong can them fat", "3"),
            ("them fat man", "4"),
            ("đầu cụt", "5"),
            ("clothed female nude male", "6"),
            ("khong can them faceless male + bald", "7"),
        ]

        for i, (note, key) in enumerate(quick_notes):
            row = i // 2
            col = i % 2

            btn = ttk.Button(
                quick_frame,
                text=f"{note} [{key}]",
                command=lambda n=note: self.add_quick_note(n)
            )

            # Cho nút 7 nằm riêng 1 hàng, span 2 cột cho dễ bấm/đọc
            if key == "7":
                btn.grid(row=row, column=0, columnspan=2, padx=2, pady=2, sticky="ew")
            else:
                btn.grid(row=row, column=col, padx=2, pady=2, sticky="ew")

        quick_frame.columnconfigure(0, weight=1)
        quick_frame.columnconfigure(1, weight=1)

        ttk.Label(
            col3_frame,
            text="Ghi chu cua ban:",
            font=('Arial', 9, 'bold')
        ).pack(anchor=tk.W, pady=(5, 0))

        self.note_text = tk.Text(
            col3_frame,
            height=8,
            width=40,
            wrap=tk.WORD,
            relief=tk.SOLID,
            borderwidth=1
        )
        self.note_text.pack(fill=tk.BOTH, pady=5)

        info_text1 = "Auto-save khi chuyen anh (Left/Right)"
        info_text2 = "Xoa note: Delete key"
        info_label = ttk.Label(
            col3_frame,
            text=info_text1 + " | " + info_text2,
            font=('Arial', 8),
            foreground='gray'
        )
        info_label.pack(pady=5)

        btn_frame = ttk.Frame(col3_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="Clear Note (Del)", command=self.clear_note).pack(fill=tk.X)

        self.main_container.columnconfigure(0, weight=1)
        self.main_container.columnconfigure(1, weight=2)
        self.main_container.columnconfigure(2, weight=1)

    # ===================== GALLERY VIEW =====================
    def create_gallery_view(self):
        for widget in self.main_container.winfo_children():
            widget.destroy()

        self.gallery_thumbnails = []

        canvas_frame = ttk.Frame(self.main_container)
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.gallery_canvas = tk.Canvas(
            canvas_frame,
            yscrollcommand=scrollbar.set,
            bg='#e0e0e0'
        )
        self.gallery_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar.config(command=self.gallery_canvas.yview)

        self.gallery_frame = ttk.Frame(self.gallery_canvas)
        self.gallery_canvas.create_window((0, 0), window=self.gallery_frame, anchor='nw')

        if not self.data_list:
            ttk.Label(
                self.gallery_frame,
                text="Chua co du lieu. Vui long Load Data truoc.",
                font=('Arial', 14)
            ).pack(pady=50)
            return

        thumb_size = 180
        cols = 5

        for idx, data in enumerate(self.data_list):
            row = idx // cols
            col = idx % cols

            item_frame = ttk.Frame(self.gallery_frame, relief=tk.SOLID, borderwidth=2, padding=5)
            item_frame.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")

            f1_canvas = tk.Canvas(
                item_frame,
                width=thumb_size,
                height=thumb_size,
                bg='#f0f0f0',
                relief=tk.SOLID,
                borderwidth=1
            )
            f1_canvas.pack()

            ttk.Label(
                item_frame,
                text=f"F1: {data['img1_name'][:20]}",
                font=('Arial', 8)
            ).pack()

            self.display_gallery_thumb(f1_canvas, data['img1'], thumb_size)

            num_f2 = len(data['img2_list'])
            ttk.Label(
                item_frame,
                text=f"F2: {num_f2} anh",
                foreground='darkgreen',
                font=('Arial', 8, 'bold')
            ).pack()

            note_display = data['note'] if data['note'] else "ok"
            note_preview = note_display[:30] + "..." if len(note_display) > 30 else note_display
            note_label = ttk.Label(
                item_frame,
                text=note_preview,
                foreground='gray',
                font=('Arial', 7)
            )
            note_label.pack()

            def make_click_handler(index):
                return lambda e: self.jump_to_image(index)

            item_frame.bind("<Button-1>", make_click_handler(idx))
            f1_canvas.bind("<Button-1>", make_click_handler(idx))

            def on_enter(e, frame=item_frame):
                frame.config(relief=tk.RAISED)

            def on_leave(e, frame=item_frame):
                frame.config(relief=tk.SOLID)

            item_frame.bind("<Enter>", on_enter)
            item_frame.bind("<Leave>", on_leave)

        self.gallery_frame.update_idletasks()
        self.gallery_canvas.config(scrollregion=self.gallery_canvas.bbox("all"))
        self.gallery_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_mousewheel(self, event):
        if self.view_mode.get() == 'gallery':
            self.gallery_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def display_gallery_thumb(self, canvas, img_path, size):
        canvas.delete("all")

        if not img_path or not os.path.exists(img_path):
            center = size // 2
            canvas.create_text(center, center, text="No image", font=('Arial', 10), fill='gray')
            return

        try:
            img = Image.open(img_path)
            img.thumbnail((size - 10, size - 10), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            x = (size - img.width) // 2
            y = (size - img.height) // 2

            canvas.create_image(x, y, anchor=tk.NW, image=photo)
            canvas.image = photo
        except:
            center = size // 2
            canvas.create_text(center, center, text="Error", font=('Arial', 10), fill='red')

    # ===================== MISC HELPERS =====================
    def jump_to_image(self, index):
        self.current_index = index
        self.switch_view('single')
        self.display_current()

    def add_quick_note(self, note_text):
        current = self.note_text.get(1.0, tk.END).strip()
        if current:
            self.note_text.insert(tk.END, ", " + note_text)
        else:
            self.note_text.insert(1.0, note_text)

    def clear_note(self):
        self.note_text.delete(1.0, tk.END)

    def auto_save_note(self):
        if self.data_list and self.view_mode.get() == 'single':
            note_content = self.note_text.get(1.0, tk.END).strip()
            self.data_list[self.current_index]['note'] = note_content

    def browse_folder(self, folder_num):
        folder = filedialog.askdirectory()
        if folder:
            if folder_num == 1:
                self.folder1_path.set(folder)
            else:
                self.folder2_path.set(folder)
            self.save_config()

    # ===================== TXT LOADING =====================
    def load_txt_content(self, folder1):
        txt_mode = self.txt_mode.get()
        txt_file_choice = self.txt_file_choice.get()

        if txt_mode == 'all_tags':
            out_tags_folder = os.path.join(folder1, 'out_tags')
            txt_filename = f"{txt_file_choice}.txt"
            txt_file_path = os.path.join(out_tags_folder, txt_filename)

            if os.path.exists(txt_file_path):
                with open(txt_file_path, 'r', encoding='utf-8') as f:
                    self.all_tags_content = [line.strip() for line in f.readlines()]
                status_msg = f"Mode: {txt_filename} ({len(self.all_tags_content)} dòng)"
                self.status_label.config(text=status_msg, foreground='blue')
                return True
            else:
                messagebox.showerror(
                    "Lỗi",
                    f"Không tìm thấy: {txt_file_path}\n\nHãy kiểm tra folder out_tags"
                )
                return False
        else:
            self.status_label.config(
                text="Mode: Anh-TXT (mỗi ảnh có 1 file .txt cùng tên)",
                foreground='blue'
            )
            return True

    # ===================== DATA LOAD =====================
    def load_data(self):
        folder1 = self.folder1_path.get()
        folder2 = self.folder2_path.get()

        if not folder1 or not folder2:
            messagebox.showerror("Lỗi", "Vui lòng chọn cả 2 folder!")
            return

        self.save_config()

        if not self.load_txt_content(folder1):
            return

        self.data_list = []
        num_per_row = self.images_per_row.get()

        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp']
        folder1_images = sorted([
            f for f in os.listdir(folder1)
            if os.path.splitext(f.lower())[1] in image_extensions
        ])
        folder2_images = sorted([
            f for f in os.listdir(folder2)
            if os.path.splitext(f.lower())[1] in image_extensions
        ])

        folder2_index = 0

        for idx, img1 in enumerate(folder1_images):
            img1_path = os.path.join(folder1, img1)

            img2_list = []
            for _ in range(num_per_row):
                if folder2_index < len(folder2_images):
                    img2_path = os.path.join(folder2, folder2_images[folder2_index])
                    img2_list.append({
                        'path': img2_path,
                        'name': folder2_images[folder2_index]
                    })
                    folder2_index += 1
                else:
                    img2_list.append({'path': None, 'name': 'N/A'})

            prompt = ""
            if self.txt_mode.get() == 'individual':
                txt_path = os.path.join(folder1, os.path.splitext(img1)[0] + '.txt')
                if os.path.exists(txt_path):
                    with open(txt_path, 'r', encoding='utf-8') as f:
                        prompt = f.read()
            else:
                if idx < len(self.all_tags_content):
                    prompt = self.all_tags_content[idx]

            self.data_list.append({
                'img1': img1_path,
                'img2_list': img2_list,
                'prompt': prompt,
                'note': '',
                'img1_name': img1
            })

        if self.data_list:
            self.current_index = 0

            if self.view_mode.get() == 'single':
                self.display_current()
            else:
                self.create_gallery_view()

            msg = f"Đã load {len(self.data_list)} ảnh từ F1, {num_per_row} ảnh/hàng từ F2"
            messagebox.showinfo("Thành công", msg)
        else:
            messagebox.showwarning("Cảnh báo", "Không tìm thấy ảnh!")

    # ===================== DISPLAY CURRENT =====================
    def display_image_on_canvas(self, canvas, img_path, canvas_size):
        canvas.delete("all")

        if not img_path or not os.path.exists(img_path):
            center = canvas_size // 2
            canvas.create_text(
                center,
                center,
                text="Không có ảnh",
                font=('Arial', 12),
                fill='gray'
            )
            return

        try:
            img = Image.open(img_path)
            max_size = canvas_size - 10
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            x = (canvas_size - img.width) // 2
            y = (canvas_size - img.height) // 2

            canvas.create_image(x, y, anchor=tk.NW, image=photo)
            canvas.image = photo
        except Exception:
            center = canvas_size // 2
            canvas.create_text(
                center,
                center,
                text="Lỗi",
                font=('Arial', 10),
                fill='red'
            )

    def display_current(self):
        if not self.data_list or self.view_mode.get() != 'single':
            return

        data = self.data_list[self.current_index]

        num_images = self.images_per_row.get()
        if num_images == 1:
            canvas_size = 400
        elif num_images == 2:
            canvas_size = 350
        else:
            canvas_size = 280

        self.display_image_on_canvas(self.canvas1, data['img1'], canvas_size)
        self.filename1_label.config(text=f"File: {data['img1_name']}")

        for i, canvas_info in enumerate(self.canvas2_list):
            if i < len(data['img2_list']):
                img2_data = data['img2_list'][i]
                self.display_image_on_canvas(canvas_info['canvas'], img2_data['path'], canvas_size)
                canvas_info['label'].config(text=img2_data['name'])
            else:
                canvas_info['canvas'].delete("all")
                canvas_info['label'].config(text="N/A")

        self.prompt_text.delete(1.0, tk.END)
        self.prompt_text.insert(1.0, data['prompt'])

        self.note_text.delete(1.0, tk.END)
        self.note_text.insert(1.0, data['note'])

        counter_text = f"{self.current_index + 1} / {len(self.data_list)}"
        self.counter_label.config(text=counter_text)

    # ===================== NAVIGATION =====================
    def previous_image(self):
        if self.data_list and self.current_index > 0:
            self.auto_save_note()
            self.current_index -= 1
            self.display_current()
            return 'break'

    def next_image(self):
        if self.data_list and self.current_index < len(self.data_list) - 1:
            self.auto_save_note()
            self.current_index += 1
            self.display_current()
            return 'break'

    # ===================== EXPORT EXCEL =====================
    def export_excel(self):
        if not self.data_list:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu!")
            return

        self.auto_save_note()

        current_dir = os.path.dirname(os.path.abspath(__file__))
        output_folder = os.path.join(current_dir, 'excel_output')
        os.makedirs(output_folder, exist_ok=True)

        folder1 = self.folder1_path.get().rstrip(r"\\/")
        base_name = os.path.basename(folder1) or "folder1"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_name_only = f"{base_name}_notes_{timestamp}.xlsx"
        filename = os.path.join(output_folder, file_name_only)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Image Comparison"

        folder1 = self.folder1_path.get()
        ws['A1'] = "Folder:"
        ws['B1'] = folder1
        ws.merge_cells('B1:E1')

        cell_a1 = ws['A1']
        cell_a1.font = openpyxl.styles.Font(bold=True, size=11)
        cell_a1.fill = openpyxl.styles.PatternFill(
            start_color="FFE699",
            end_color="FFE699",
            fill_type="solid"
        )

        cell_b1 = ws['B1']
        cell_b1.font = openpyxl.styles.Font(size=11, color="0000FF")
        cell_b1.alignment = openpyxl.styles.Alignment(horizontal='left')

        ws.row_dimensions[1].height = 25

        num_img2 = self.images_per_row.get()
        headers = ["STT", "Image 1"]
        for i in range(num_img2):
            headers.append(f"Image 2-{i+1}")
        headers.extend(["Prompt", "Notes"])

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="CCE5FF",
                end_color="CCE5FF",
                fill_type="solid"
            )

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        for i in range(num_img2):
            ws.column_dimensions[chr(67 + i)].width = 30
        ws.column_dimensions[chr(67 + num_img2)].width = 50
        ws.column_dimensions[chr(68 + num_img2)].width = 50

        for idx, data in enumerate(self.data_list, start=4):
            ws[f"A{idx}"] = idx - 3
            ws[f"B{idx}"] = data['img1_name']

            for i, img2 in enumerate(data['img2_list']):
                col_letter = chr(67 + i)
                ws[f"{col_letter}{idx}"] = img2['name']

            prompt_col = chr(67 + num_img2)
            note_col = chr(68 + num_img2)
            ws[f"{prompt_col}{idx}"] = data['prompt']

            note_value = data['note'] if data['note'] else "ok"
            ws[f"{note_col}{idx}"] = note_value

            ws.row_dimensions[idx].height = 30

        ws_info = wb.create_sheet("Info")
        ws_info['A1'] = "Folder 1:"
        ws_info['B1'] = self.folder1_path.get()
        ws_info['A2'] = "Folder 2:"
        ws_info['B2'] = self.folder2_path.get()
        ws_info['A3'] = "TXT Mode:"
        txt_mode_display = "Anh-TXT" if self.txt_mode.get() == 'individual' else "out_tags folder"
        ws_info['B3'] = txt_mode_display
        ws_info['A4'] = "TXT File Choice:"
        ws_info['B4'] = self.txt_file_choice.get()
        ws_info['A5'] = "Images per row:"
        ws_info['B5'] = self.images_per_row.get()
        ws_info['A6'] = "Total F1 Images:"
        ws_info['B6'] = len(self.data_list)
        ws_info['A7'] = "Export Date:"
        ws_info['B7'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_info['A8'] = "Excel Output Folder:"
        ws_info['B8'] = output_folder

        wb.save(filename)

        messagebox.showinfo("Thành công", f"Đã xuất: {os.path.basename(filename)}")
        os.startfile(output_folder)


if __name__ == "__main__":
    root = tk.Tk()
    app = ImageCompareApp(root)
    root.mainloop()
