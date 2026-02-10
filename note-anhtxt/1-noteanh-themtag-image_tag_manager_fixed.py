import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox
from PIL import Image, ImageTk
import os
import re
import threading
import openpyxl
import json  # Thêm thư viện để lưu cấu hình

# Khi Notes đã xử lý xong thì ghi gì vào Excel?
CLEAR_NOTES_TO = ""  # hoặc "ok"

# Thời gian debounce lưu Excel (ms)
DEBOUNCE_MS = 700

# Danh sách tag cố định (key để hiển thị, value chèn vào prompt)
REQUIRED_TAGS = {
    "fat man": "fat man",
    "bald": "bald",
    "faceless male": "faceless male",
}

# --- CẤU HÌNH NHỚ FOLDER ---
CONFIG_FILE = "config.json"
DEFAULT_DIR = r"D:\no\note-anhtxt\excel_output"


class ImageTagManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Image Tag Manager - Optimized Excel Notes")
        self.root.geometry("1400x850")

        # excel
        self.excel_file = ""
        self.sheet_name = ""
        self.wb = None
        self.ws = None

        self.header_row = None
        self.stt_col = 0
        self.image_col = -1
        self.prompt_col = -1
        self.notes_col = -1

        # data
        self.folder_path = ""
        self.excel_data = []
        self.current_index = 0

        # UI state
        self.tags_buttons = []   # list of (tag_value, button)
        self.thumb_items = []    # {"frame","idx","widgets_to_recolor","badge","lb3"}
        self._mousewheel_bound = False

        # save performance (debounce + thread)
        self._dirty = False
        self._saving = False
        self._save_job = None

        # danh sách tag yêu cầu
        self.required_tags = REQUIRED_TAGS.copy()
        
        # Folder ghi nhớ
        self.last_dir = DEFAULT_DIR
        self.load_last_dir()

        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # ----------------- CONFIG / MEMORY -----------------

    def load_last_dir(self):
        """Đọc folder cũ từ file json, nếu không có thì dùng mặc định"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    saved_dir = data.get("last_dir", "")
                    if saved_dir and os.path.exists(saved_dir):
                        self.last_dir = saved_dir
            except Exception:
                pass  # Lỗi thì cứ dùng default
    
    def save_last_dir(self):
        """Lưu folder hiện tại vào json"""
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({"last_dir": self.last_dir}, f)
        except Exception:
            pass

    # ----------------- BASIC UTILS -----------------

    def natural_sort_key(self, s: str):
        return [int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", s)]

    def list_images_in_folder(self, folder):
        exts = (".png", ".jpg", ".jpeg", ".webp", ".bmp")
        files = []
        if os.path.exists(folder):
            for name in os.listdir(folder):
                p = os.path.join(folder, name)
                if os.path.isfile(p) and name.lower().endswith(exts):
                    files.append(name)
        files.sort(key=self.natural_sort_key)
        return files

    def coerce_numeric_string(self, raw):
        if raw is None:
            return ""
        s = str(raw).strip()
        if not s:
            return ""
        if re.fullmatch(r"\d+(\.0+)?", s):
            try:
                return str(int(float(s)))
            except Exception:
                return s
        return s

    def resolve_image_name(self, raw_cell_value, folder_images, seq_index_zero_based):
        raw_s = self.coerce_numeric_string(raw_cell_value)

        # filename có đuôi
        if raw_s and "." in raw_s:
            return raw_s

        # số thứ tự 1-based
        if raw_s and re.fullmatch(r"\d+", raw_s):
            idx = int(raw_s) - 1
            if 0 <= idx < len(folder_images):
                return folder_images[idx]
            if 0 <= seq_index_zero_based < len(folder_images):
                return folder_images[seq_index_zero_based]
            return raw_s

        # fallback theo thứ tự dòng
        if 0 <= seq_index_zero_based < len(folder_images):
            return folder_images[seq_index_zero_based]

        return raw_s

    # ----------------- NOTES -----------------

    def notes_is_warning(self, notes: str) -> bool:
        n = (notes or "").strip().lower()
        return bool(n) and (n != "ok")

    def parse_notes_and_remove_done(self, notes: str, prompt_text: str) -> str:
        """
        Notes dạng: 'thêm fat man, thiếu bald'
        - Nếu prompt đã có tag tương ứng => remove item đó khỏi Notes.
        - Nếu item không đúng format => giữ nguyên.
        """
        n = (notes or "").strip()
        if not n or n.lower() == "ok":
            return n

        prompt_lc = (prompt_text or "").lower()

        parts = [p.strip() for p in n.split(",") if p.strip()]
        keep = []

        for p in parts:
            p_lc = p.lower().strip()
            m = re.match(r"^(thêm|them|thiếu|thieu)\s+(.+)$", p_lc)
            if not m:
                keep.append(p.strip())
                continue

            tag = m.group(2).strip()
            if not tag:
                keep.append(p.strip())
                continue

            if tag in prompt_lc:
                # tag đã có trong prompt => bỏ item này
                continue

            keep.append(p.strip())

        return ", ".join(keep).strip()

    def extract_tags_from_notes(self, notes: str):
        """
        Trích ra danh sách tag từ Notes dạng 'thêm xx', 'thiếu yy'
        để quyết định button nào cần tô đỏ.
        """
        n = (notes or "").strip()
        if not n or n.lower() == "ok":
            return set()

        parts = [p.strip() for p in n.split(",") if p.strip()]
        tags = set()
        for p in parts:
            p_lc = p.lower()
            m = re.match(r"^(thêm|them|thiếu|thieu)\s+(.+)$", p_lc)
            if not m:
                continue
            tag = m.group(2).strip()
            if tag:
                tags.add(tag)
        return tags

    # ----------------- EXCEL: KEEP OPEN -----------------

    def close_workbook(self):
        try:
            if self.wb:
                self.wb.close()
        except Exception:
            pass
        self.wb = None
        self.ws = None

    def find_header_row_and_folder(self, ws):
        folder_path = ""
        header_row = None

        # scan nhanh 1..120 dòng đầu
        max_scan = min(120, ws.max_row)
        for r in range(1, max_scan + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value

            if a is not None and str(a).strip() == "Folder:" and b:
                folder_path = str(b).strip()

            if a is not None and str(a).strip().lower() == "stt":
                header_row = r
                break

        return folder_path, header_row

    def map_columns(self, ws, header_row):
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            headers.append(str(v).strip() if v is not None else "")
        headers_lc = [h.lower() for h in headers]

        def find_exact(name):
            name_lc = name.lower()
            for i, h in enumerate(headers_lc):
                if h == name_lc:
                    return i
            return -1

        stt_col = find_exact("stt")
        if stt_col < 0:
            stt_col = 0

        prompt_col = find_exact("prompt")
        notes_col = find_exact("notes")

        image_col = -1
        for i, h in enumerate(headers_lc):
            if h.startswith("image"):
                image_col = i
                break
        if image_col < 0:
            for i, h in enumerate(headers_lc):
                if "image" in h:
                    image_col = i
                    break

        return stt_col, image_col, prompt_col, notes_col

    def set_excel_cell(self, excel_row: int, col_index0: int, value):
        # chỉ set cell trong workbook đang mở (chưa save ngay)
        if not self.ws or excel_row is None or col_index0 < 0:
            return False
        try:
            self.ws.cell(row=excel_row, column=col_index0 + 1).value = value
            return True
        except Exception:
            return False

    # ----------------- SAVE: DEBOUNCE + THREAD -----------------

    def mark_dirty_and_schedule_save(self):
        self._dirty = True
        if self._save_job is not None:
            try:
                self.root.after_cancel(self._save_job)
            except Exception:
                pass
        # dùng hằng DEBOUNCE_MS
        self._save_job = self.root.after(DEBOUNCE_MS, self.flush_save_async)

    def flush_save_async(self):
        self._save_job = None
        if not self._dirty:
            return
        if self._saving:
            # đang save, cứ để _dirty=True, thread xong sẽ gọi lại
            return
        if not self.wb or not self.excel_file:
            return

        self._saving = True
        self._dirty = False
        self.status_label.config(text="💾 Đang lưu Excel...")

        def worker():
            err = None
            try:
                # Workbook.save sẽ ghi đè file; nếu đang bị Excel giữ lock sẽ ném exception nên cần try/except
                self.wb.save(self.excel_file)
            except Exception as e:
                err = str(e)

            def on_done():
                self._saving = False
                if err:
                    self.status_label.config(text=f"❌ Lưu Excel lỗi (có thể file đang mở): {err[:80]}")
                else:
                    self.status_label.config(text="✅ Đã lưu Excel")

                # nếu trong lúc save lại có thay đổi => save tiếp
                if self._dirty:
                    self.mark_dirty_and_schedule_save()

            self.root.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()

    # ----------------- UI -----------------

    def setup_ui(self):
        top_frame = tk.LabelFrame(self.root, text="📊 CHỌN FILE EXCEL", font=("Arial", 11, "bold"), bg="#e3f2fd")
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        excel_frame = tk.Frame(top_frame, bg="#e3f2fd")
        excel_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Button(
            excel_frame, text="📊 Chọn file Excel",
            command=self.load_from_excel,
            bg="#FF9800", fg="white",
            font=("Arial", 11, "bold"), width=20
        ).pack(side=tk.LEFT, padx=5)

        self.excel_path_label = tk.Label(excel_frame, text="Chưa chọn file Excel", font=("Arial", 10), bg="#e3f2fd", fg="gray")
        self.excel_path_label.pack(side=tk.LEFT, padx=10)

        self.folder_label = tk.Label(top_frame, text="", font=("Arial", 9), bg="#e3f2fd", fg="blue")
        self.folder_label.pack(fill=tk.X, padx=10, pady=5)

        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # LEFT thumbnails
        left_panel = tk.Frame(main_frame, width=260, relief=tk.RIDGE, bd=2)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=5)
        left_panel.pack_propagate(False)

        tk.Label(left_panel, text="📁 DANH SÁCH ẢNH", font=("Arial", 12, "bold"), bg="#e3f2fd").pack(fill=tk.X, pady=5)

        canvas_frame = tk.Frame(left_panel)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.thumb_canvas = tk.Canvas(canvas_frame, bg="white")
        scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=self.thumb_canvas.yview)
        self.thumb_frame = tk.Frame(self.thumb_canvas, bg="white")

        self.thumb_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.thumb_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.thumb_canvas.create_window((0, 0), window=self.thumb_frame, anchor="nw")
        self.thumb_frame.bind("<Configure>", lambda e: self.thumb_canvas.configure(scrollregion=self.thumb_canvas.bbox("all")))

        self.thumb_canvas.bind("<Enter>", self._bind_mousewheel)
        self.thumb_canvas.bind("<Leave>", self._unbind_mousewheel)

        # CENTER image
        center_panel = tk.Frame(main_frame, width=600, relief=tk.RIDGE, bd=2)
        center_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        tk.Label(center_panel, text="🖼️ ẢNH ĐANG CHỌN", font=("Arial", 12, "bold"), bg="#fff3e0").pack(fill=tk.X, pady=5)

        self.filename_label = tk.Label(center_panel, text="", font=("Arial", 10), fg="blue")
        self.filename_label.pack(pady=2)

        self.notes_label = tk.Label(center_panel, text="", font=("Arial", 9, "italic"), fg="red", wraplength=550)
        self.notes_label.pack(pady=2)

        self.image_border_frame = tk.Frame(center_panel, bg="white", relief=tk.SOLID, bd=5)
        self.image_border_frame.pack(pady=10)

        self.image_canvas = tk.Canvas(self.image_border_frame, bg="gray", width=570, height=500)
        self.image_canvas.pack()

        nav_frame = tk.Frame(center_panel)
        nav_frame.pack(pady=10)

        tk.Button(nav_frame, text="◀ Trước (←)", command=self.prev_image,
                  bg="#2196F3", fg="white", font=("Arial", 11, "bold"), width=15).pack(side=tk.LEFT, padx=10)

        self.index_label = tk.Label(nav_frame, text="0/0", font=("Arial", 11, "bold"))
        self.index_label.pack(side=tk.LEFT, padx=10)

        tk.Button(nav_frame, text="Sau (→) ▶", command=self.next_image,
                  bg="#2196F3", fg="white", font=("Arial", 11, "bold"), width=15).pack(side=tk.LEFT, padx=10)

        # RIGHT editor
        right_panel = tk.Frame(main_frame, width=450, relief=tk.RIDGE, bd=2)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=5)
        right_panel.pack_propagate(False)

        tk.Label(right_panel, text="📝 PROMPT EDITOR", font=("Arial", 12, "bold"), bg="#f3e5f5").pack(fill=tk.X, pady=5)

        text_frame = tk.Frame(right_panel)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.text_area = scrolledtext.ScrolledText(text_frame, wrap=tk.WORD, font=("Arial", 10), height=12)
        self.text_area.pack(fill=tk.BOTH, expand=True)

        tags_frame = tk.LabelFrame(right_panel, text="⚡ THÊM TAGS (1-9)", font=("Arial", 10, "bold"), bg="#e8f5e9")
        tags_frame.pack(fill=tk.X, padx=5, pady=10)

        # buttons tag + hotkey 1-9
        for i, (tag_key, tag_value) in enumerate(self.required_tags.items(), 1):
            btn = tk.Button(
                tags_frame,
                text=f"{i}. {tag_key.upper()}",
                command=lambda t=tag_value: self.toggle_tag(t),
                font=("Arial", 10, "bold"),
                bg="#4CAF50",
                fg="white",
                relief=tk.RAISED,
                bd=3
            )
            btn.pack(fill=tk.X, padx=5, pady=3)
            self.tags_buttons.append((tag_value, btn))
            # bind phím số
            if i <= 9:
                self.root.bind(str(i), lambda e, t=tag_value: self.toggle_tag(t))

        tk.Button(
            right_panel,
            text="💾 SAVE (Ctrl+S) + CLEAR NOTES + SAVE EXCEL (DEBOUNCE)",
            command=self.save_prompt,
            bg="#FF9800",
            fg="white",
            font=("Arial", 11, "bold"),
            height=2,
        ).pack(fill=tk.X, padx=5, pady=5)

        self.root.bind("<Left>", lambda e: self.prev_image())
        self.root.bind("<Right>", lambda e: self.next_image())
        self.root.bind("<Control-s>", lambda e: self.save_prompt())

        self.status_label = tk.Label(self.root, text="Sẵn sàng...", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    # ----------------- MOUSE WHEEL -----------------

    def _on_mousewheel(self, event):
        if getattr(event, "num", None) == 4:
            delta = -1
        elif getattr(event, "num", None) == 5:
            delta = 1
        else:
            delta = int(-1 * (event.delta / 120))
            if delta == 0:
                delta = -1 if event.delta > 0 else 1
        self.thumb_canvas.yview_scroll(delta, "units")

    def _bind_mousewheel(self, event=None):
        if self._mousewheel_bound:
            return
        self._mousewheel_bound = True
        self.root.bind_all("<MouseWheel>", self._on_mousewheel)
        self.root.bind_all("<Button-4>", self._on_mousewheel)
        self.root.bind_all("<Button-5>", self._on_mousewheel)

    def _unbind_mousewheel(self, event=None):
        if not self._mousewheel_bound:
            return
        self._mousewheel_bound = False
        self.root.unbind_all("<MouseWheel>")
        self.root.unbind_all("<Button-4>")
        self.root.unbind_all("<Button-5>")

    # ----------------- LOAD EXCEL -----------------

    def load_from_excel(self):
        # Kiểm tra folder cũ có tồn tại không, nếu không thì dùng mặc định hoặc thư mục gốc
        init_dir = self.last_dir if os.path.exists(self.last_dir) else "/"
        
        excel_file = filedialog.askopenfilename(
            title="Chọn file Excel",
            initialdir=init_dir, # Dùng folder đã nhớ
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not excel_file:
            return
            
        # -- LƯU LẠI FOLDER MỚI NẾU KHÁC FOLDER CŨ --
        new_dir = os.path.dirname(excel_file)
        if new_dir and new_dir != self.last_dir:
            self.last_dir = new_dir
            self.save_last_dir() # Lưu vào config.json

        # đóng workbook cũ nếu có
        self.close_workbook()

        try:
            self.excel_file = excel_file
            # giữ workbook open trong RAM
            self.wb = openpyxl.load_workbook(excel_file)
            self.ws = self.wb.active
            self.sheet_name = self.ws.title

            folder_path, header_row = self.find_header_row_and_folder(self.ws)
            if not header_row:
                messagebox.showerror("Lỗi", "Không tìm thấy header dòng 'STT'!")
                return
            if not folder_path:
                messagebox.showerror("Lỗi", "Không tìm thấy dòng 'Folder:'!")
                return
            if not os.path.exists(folder_path):
                messagebox.showerror("Lỗi", f"Folder KHÔNG tồn tại:\n{folder_path}")
                return

            self.folder_path = folder_path
            self.header_row = header_row
            self.stt_col, self.image_col, self.prompt_col, self.notes_col = self.map_columns(self.ws, header_row)

            # debug cột Notes / sheet
            print(f"[DEBUG] Active sheet       : {self.sheet_name}")
            print(f"[DEBUG] Header row         : {self.header_row}")
            print(f"[DEBUG] STT col           : {self.stt_col}")
            print(f"[DEBUG] Image col         : {self.image_col}")
            print(f"[DEBUG] Prompt col        : {self.prompt_col}")
            print(f"[DEBUG] Notes col (0-based): {self.notes_col}")

            if self.image_col < 0 or self.notes_col < 0:
                messagebox.showerror("Lỗi", "Thiếu cột Image hoặc Notes trong header!")
                return

            folder_images = self.list_images_in_folder(folder_path)
            if not folder_images:
                messagebox.showerror("Lỗi", f"Folder không có ảnh:\n{folder_path}")
                return

            self.excel_data = []
            seq = 0

            # đọc dữ liệu từ worksheet đang mở
            for r in range(header_row + 1, self.ws.max_row + 1):
                stt = self.ws.cell(row=r, column=self.stt_col + 1).value
                if not stt:
                    continue

                raw_img = self.ws.cell(row=r, column=self.image_col + 1).value
                img_name = self.resolve_image_name(raw_img, folder_images, seq)

                prompt_val = ""
                notes_val = ""
                if self.prompt_col >= 0:
                    v = self.ws.cell(row=r, column=self.prompt_col + 1).value
                    if v:
                        prompt_val = str(v).strip()
                v = self.ws.cell(row=r, column=self.notes_col + 1).value
                if v:
                    notes_val = str(v).strip()

                self.excel_data.append({
                    "stt": stt,
                    "image": img_name,
                    "prompt": prompt_val,
                    "notes": notes_val,
                    "excel_row": r,
                })
                # debug vài dòng notes đầu
                if len(self.excel_data) <= 3:
                    print(f"[DEBUG] Row {r} - image={img_name} - notes='{notes_val}'")
                seq += 1

            self.excel_path_label.config(text="✓ " + os.path.basename(excel_file), fg="green")
            self.folder_label.config(text=f"📂 {folder_path} | Sheet: {self.sheet_name}")

            self.current_index = 0
            self.create_thumbnails()
            if self.excel_data:
                self.display_current_image()

            self.status_label.config(text=f"✅ Load {len(self.excel_data)} dòng (Workbook giữ mở)")

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("Lỗi", f"Lỗi load Excel:\n{str(e)}")

    # ----------------- THUMBNAILS -----------------

    def _bind_click_select(self, widget, idx):
        widget.bind("<Button-1>", lambda e, i=idx: self.select_image(i))

    def scroll_thumbnail_to_current(self):
        if not self.thumb_items:
            return
        self.root.update_idletasks()
        bbox = self.thumb_canvas.bbox("all")
        if not bbox:
            return

        total_h = max(1, bbox[3] - bbox[1])
        view_h = max(1, self.thumb_canvas.winfo_height())

        frame = self.thumb_items[self.current_index]["frame"]
        y = frame.winfo_y()

        target = max(0, y - int(view_h * 0.35))
        frac = min(1.0, max(0.0, target / total_h))
        self.thumb_canvas.yview_moveto(frac)

    def update_thumbnail_item(self, idx: int):
        """Chỉ update 1 thumbnail (notes/badge/lb3 + nền)"""
        if idx < 0 or idx >= len(self.thumb_items):
            return

        item = self.thumb_items[idx]
        notes = str(self.excel_data[idx].get("notes") or "").strip()
        warning = self.notes_is_warning(notes)

        base_bg = "#ffecec" if warning else "white"

        # show/hide badge + lb3
        badge = item.get("badge")
        lb3 = item.get("lb3")

        if warning:
            if badge and not badge.winfo_ismapped():
                badge.pack(side=tk.RIGHT, padx=6, pady=6)
            if lb3:
                if not lb3.winfo_ismapped():
                    lb3.pack(anchor="w")
                lb3.config(text=f"⚠ {notes[:28]}", fg="red")
        else:
            if badge and badge.winfo_ismapped():
                badge.pack_forget()
            if lb3 and lb3.winfo_ismapped():
                lb3.pack_forget()

        # cập nhật nền cơ bản; highlight sẽ do update_thumbnail_highlight() xử lý sau
        item["frame"].config(bg=base_bg)
        for w in item["widgets_to_recolor"]:
            try:
                w.config(bg=base_bg)
            except Exception:
                pass

    def update_thumbnail_highlight(self):
        for item in self.thumb_items:
            idx = item["idx"]
            notes = str(self.excel_data[idx].get("notes") or "").strip()
            warning = self.notes_is_warning(notes)

            base_bg = "#ffecec" if warning else "white"
            if idx == self.current_index:
                frame_bg = "#c8e6c9"
                item["frame"].config(bg=frame_bg, bd=4, relief=tk.SOLID)
            else:
                frame_bg = base_bg
                item["frame"].config(bg=frame_bg, bd=2, relief=tk.RAISED)

            for w in item["widgets_to_recolor"]:
                try:
                    w.config(bg=frame_bg)
                except Exception:
                    pass

            if item.get("badge") is not None:
                item["badge"].config(bg="#e53935", fg="white")

    def create_thumbnails(self):
        for w in self.thumb_frame.winfo_children():
            w.destroy()

        self.thumb_items = []

        for i, data in enumerate(self.excel_data):
            notes = str(data.get("notes") or "").strip()
            warning = self.notes_is_warning(notes)
            base_bg = "#ffecec" if warning else "white"

            frame = tk.Frame(self.thumb_frame, bg=base_bg, relief=tk.RAISED, bd=2)
            frame.pack(fill=tk.X, padx=5, pady=3)

            widgets_to_recolor = [frame]

            # badge luôn tạo, lúc warning mới pack
            badge = tk.Label(frame, text="⚠", fg="white", bg="#e53935", font=("Arial", 10, "bold"), width=2)
            if warning:
                badge.pack(side=tk.RIGHT, padx=6, pady=6)

            img_path = os.path.join(self.folder_path, data["image"])

            # ảnh thumbnail
            if os.path.exists(img_path):
                try:
                    img = Image.open(img_path)
                    img.thumbnail((160, 160))
                    photo = ImageTk.PhotoImage(img)
                    label_img = tk.Label(frame, image=photo, bg=base_bg)
                    label_img.image = photo
                    label_img.pack(side=tk.LEFT, padx=5, pady=5)
                    widgets_to_recolor.append(label_img)
                    self._bind_click_select(label_img, i)
                except Exception:
                    label_x = tk.Label(frame, text="❌", font=("Arial", 26), fg="red", bg=base_bg)
                    label_x.pack(side=tk.LEFT, padx=20, pady=5)
                    widgets_to_recolor.append(label_x)
                    self._bind_click_select(label_x, i)
            else:
                label_x = tk.Label(frame, text="❌", font=("Arial", 26), fg="red", bg=base_bg)
                label_x.pack(side=tk.LEFT, padx=20, pady=5)
                widgets_to_recolor.append(label_x)
                self._bind_click_select(label_x, i)

            text_frame = tk.Frame(frame, bg=base_bg)
            text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
            widgets_to_recolor.append(text_frame)

            lb1 = tk.Label(text_frame, text=f"#{data['stt']}", font=("Arial", 10, "bold"), bg=base_bg)
            lb1.pack(anchor="w")
            widgets_to_recolor.append(lb1)

            lb2 = tk.Label(text_frame, text=str(data["image"])[:30], font=("Arial", 8), bg=base_bg)
            lb2.pack(anchor="w")
            widgets_to_recolor.append(lb2)

            # lb3 luôn tạo, warning mới pack
            lb3 = tk.Label(text_frame, text=f"⚠ {notes[:28]}", font=("Arial", 7, "italic"), fg="red", bg=base_bg)
            if warning:
                lb3.pack(anchor="w")
            widgets_to_recolor.append(lb3)

            # click chọn
            self._bind_click_select(frame, i)
            self._bind_click_select(text_frame, i)
            self._bind_click_select(lb1, i)
            self._bind_click_select(lb2, i)
            self._bind_click_select(lb3, i)
            self._bind_click_select(badge, i)

            self.thumb_items.append({
                "frame": frame,
                "idx": i,
                "widgets_to_recolor": widgets_to_recolor,
                "badge": badge,
                "lb3": lb3,
            })

        self.update_thumbnail_highlight()
        self.scroll_thumbnail_to_current()

    # ----------------- MAIN DISPLAY -----------------

    def select_image(self, index):
        # chỉ lưu txt + update notes (excel save debounce)
        self.save_prompt(show_message=False)
        self.current_index = index
        self.display_current_image()

    def display_current_image(self):
        if not self.excel_data:
            return

        data = self.excel_data[self.current_index]
        img_path = os.path.join(self.folder_path, data["image"])
        self.filename_label.config(text=f"📄 #{data['stt']} - {data['image']}")

        notes_text = str(data.get("notes") or "").strip()
        if self.notes_is_warning(notes_text):
            self.notes_label.config(text="⚠️ " + notes_text, fg="red")
            self.image_border_frame.config(bg="#ffcccc", bd=8)
        else:
            self.notes_label.config(text="✅ OK", fg="green")
            self.image_border_frame.config(bg="white", bd=5)

        try:
            if os.path.exists(img_path):
                img = Image.open(img_path)
                img.thumbnail((570, 500))
                photo = ImageTk.PhotoImage(img)
                self.image_canvas.delete("all")
                self.image_canvas.create_image(285, 250, image=photo)
                self.image_canvas.image = photo
            else:
                self.image_canvas.delete("all")
                self.image_canvas.create_text(285, 250, text=f"❌ Missing:\n{data['image']}", font=("Arial", 12), fill="red", width=500)
        except Exception as e:
            self.image_canvas.delete("all")
            self.image_canvas.create_text(285, 250, text=f"❌ Error: {str(e)}", font=("Arial", 12), fill="red")

        self.index_label.config(text=f"{self.current_index + 1}/{len(self.excel_data)}")

        self.load_prompt()
        self.update_thumbnail_highlight()
        self.scroll_thumbnail_to_current()
        self.update_tag_buttons_for_current()

    # ----------------- PROMPT TXT + CLEAR NOTES -----------------

    def load_prompt(self):
        data = self.excel_data[self.current_index]
        img_path = os.path.join(self.folder_path, data["image"])
        txt_path = os.path.splitext(img_path)[0] + ".txt"

        prompt = data.get("prompt", "")
        if os.path.exists(txt_path):
            try:
                with open(txt_path, "r", encoding="utf-8") as f:
                    prompt = f.read().strip()
            except Exception:
                pass

        self.text_area.delete(1.0, tk.END)
        self.text_area.insert(1.0, prompt)

    def save_prompt(self, show_message=True):
        if not self.excel_data:
            return

        data = self.excel_data[self.current_index]
        img_path = os.path.join(self.folder_path, data["image"])
        txt_path = os.path.splitext(img_path)[0] + ".txt"

        content = self.text_area.get(1.0, tk.END).strip()
        self.excel_data[self.current_index]["prompt"] = content

        # save txt prompt
        try:
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write(content)
        except Exception as e:
            self.status_label.config(text="❌ Lưu txt lỗi: " + str(e)[:60])
            return

        # clear Notes in memory + set cell in workbook (NO SAVE IMMEDIATELY)
        self.auto_clear_notes_current_row()

        if show_message:
            self.status_label.config(text="✅ Saved txt (Excel save debounce...)")

    def refresh_current_notes_ui(self):
        """
        Cập nhật label Notes, viền ảnh lớn, thumbnail hiện tại và trạng thái button tag.
        Không load lại ảnh/prompt để tránh giật.
        """
        data = self.excel_data[self.current_index]
        notes_text = str(data.get("notes") or "").strip()

        if self.notes_is_warning(notes_text):
            self.notes_label.config(text="⚠️ " + notes_text, fg="red")
            self.image_border_frame.config(bg="#ffcccc", bd=8)
        else:
            self.notes_label.config(text="✅ OK", fg="green")
            self.image_border_frame.config(bg="white", bd=5)

        # update đúng 1 thumbnail
        self.update_thumbnail_item(self.current_index)
        self.update_thumbnail_highlight()
        self.update_tag_buttons_for_current()

    def auto_clear_notes_current_row(self):
        data = self.excel_data[self.current_index]
        old_notes = str(data.get("notes") or "").strip()
        if not old_notes or old_notes.lower() == "ok":
            # vẫn nên update trạng thái button (trường hợp chỉ chỉnh prompt)
            self.update_tag_buttons_for_current()
            return

        prompt_text = self.excel_data[self.current_index]["prompt"]
        new_notes = self.parse_notes_and_remove_done(old_notes, prompt_text)

        if not new_notes:
            new_notes = CLEAR_NOTES_TO

        if new_notes == old_notes:
            self.update_tag_buttons_for_current()
            return

        # update memory
        self.excel_data[self.current_index]["notes"] = new_notes

        # set excel cell in wb (still in RAM)
        ok = self.set_excel_cell(data["excel_row"], self.notes_col, new_notes)
        if ok:
            # chỉ update đúng UI hiện tại, không rebuild thumbnails
            self.refresh_current_notes_ui()
            # schedule actual save to disk (debounced + background)
            self.mark_dirty_and_schedule_save()

    # ----------------- TAGS -----------------

    def update_tag_buttons_for_current(self):
        """
        Quy tắc:
        - Nếu prompt đã chứa tag => button xanh đậm + SUNKEN.
        - Nếu Notes có yêu cầu thêm/thiếu tag đó mà prompt chưa có => button đỏ.
        - Nếu Notes không nhắc tag đó => button xanh bình thường.
        """
        if not self.excel_data or not self.tags_buttons:
            return

        prompt = self.text_area.get(1.0, tk.END).strip().lower()
        notes = str(self.excel_data[self.current_index].get("notes") or "")
        tags_need = self.extract_tags_from_notes(notes)

        for tag_value, btn in self.tags_buttons:
            tag_lc = tag_value.lower()

            if tag_lc in prompt:
                # đã có trong prompt: xanh đậm + sunken để mô phỏng trạng thái đang bật
                btn.config(bg="#1B5E20", fg="white", relief=tk.SUNKEN)
            else:
                if tag_lc in tags_need:
                    # Notes bảo thiếu/thêm tag này mà prompt chưa có => đỏ
                    btn.config(bg="#D32F2F", fg="white", relief=tk.RAISED)
                else:
                    # bình thường
                    btn.config(bg="#4CAF50", fg="white", relief=tk.RAISED)

    def toggle_tag(self, tag):
        content = self.text_area.get(1.0, tk.END).strip()
        if tag.lower() in content.lower():
            # remove tag (dùng regex word-boundary đơn giản)
            content = re.sub(rf"(?i)\b{re.escape(tag)}\b", "", content)
            content = re.sub(r",\s*,", ",", content)
            content = content.replace(",,", ",").strip(" ,")
        else:
            content = content + ", " + tag if content else tag

        self.text_area.delete(1.0, tk.END)
        self.text_area.insert(1.0, content)

        # cập nhật màu nút ngay sau khi toggle
        self.update_tag_buttons_for_current()

    # ----------------- NAV -----------------

    def prev_image(self):
        if self.excel_data:
            self.save_prompt(show_message=False)
            self.current_index = (self.current_index - 1) % len(self.excel_data)
            self.display_current_image()

    def next_image(self):
        if self.excel_data:
            self.save_prompt(show_message=False)
            self.current_index = (self.current_index + 1) % len(self.excel_data)
            self.display_current_image()

    # ----------------- CLOSE -----------------

    def on_close(self):
        # flush save lần cuối (blocking) để khỏi mất dữ liệu
        try:
            if self.wb and self.excel_file and self._dirty and (not self._saving):
                self.status_label.config(text="💾 Đang lưu Excel lần cuối...")
                self.wb.save(self.excel_file)
        except Exception:
            pass
        finally:
            self.close_workbook()
            self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = ImageTagManager(root)
    root.mainloop()
